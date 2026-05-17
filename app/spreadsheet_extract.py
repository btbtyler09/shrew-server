"""Spreadsheet (xlsx/xls/ods) extraction.

Reads workbooks deterministically via openpyxl, emitting per-sheet markdown:
- Data sheets become GFM tables (merged cells flattened by value-repetition).
- Chart sheets become text blocks describing title, type, axis labels, and series.
- Mixed sheets emit both.

Non-xlsx inputs (.xls, .ods) are converted to .xlsx via LibreOffice first.
"""

import logging
import os
import re
import subprocess
from datetime import date, datetime, time
from typing import Any

logger = logging.getLogger("shrew.spreadsheet_extract")

SPREADSHEET_MAX_ROWS = 10000

_RANGE_RE = re.compile(
    r"""
    (?:
        '(?P<quoted_sheet>(?:[^']|'')+)'
        |
        (?P<sheet>[A-Za-z_][A-Za-z0-9_.]*)
    )
    !
    \$?(?P<col1>[A-Z]+)\$?(?P<row1>\d+)
    (?: : \$?(?P<col2>[A-Z]+)\$?(?P<row2>\d+) )?
    """,
    re.VERBOSE,
)

_CHART_TYPE_MAP = {
    "BarChart": "bar",
    "BarChart3D": "bar",
    "LineChart": "line",
    "LineChart3D": "line",
    "ScatterChart": "scatter",
    "PieChart": "pie",
    "PieChart3D": "pie",
    "DoughnutChart": "doughnut",
    "AreaChart": "area",
    "AreaChart3D": "area",
    "RadarChart": "radar",
    "BubbleChart": "bubble",
    "SurfaceChart": "surface",
    "SurfaceChart3D": "surface",
    "StockChart": "stock",
}


def _convert_to_xlsx(src_path: str, output_dir: str) -> str:
    """Convert .xls / .ods to .xlsx via LibreOffice headless."""
    basename = os.path.basename(src_path)
    logger.info(f"Converting {basename} to xlsx via LibreOffice")
    try:
        result = subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "xlsx",
             "--outdir", output_dir, src_path],
            capture_output=True, text=True, timeout=120,
        )
    except FileNotFoundError:
        raise RuntimeError(
            "LibreOffice is not installed. "
            "Install it to process .xls/.ods spreadsheets: apt install libreoffice"
        )
    except subprocess.TimeoutExpired:
        raise RuntimeError(
            f"LibreOffice xlsx conversion timed out after 120s for {basename}"
        )
    if result.returncode != 0:
        raise RuntimeError(
            f"LibreOffice xlsx conversion failed for {basename}: {result.stderr}"
        )
    stem = os.path.splitext(basename)[0]
    xlsx_path = os.path.join(output_dir, f"{stem}.xlsx")
    if not os.path.exists(xlsx_path):
        raise RuntimeError(
            f"LibreOffice did not produce expected xlsx at {xlsx_path}"
        )
    return xlsx_path


def _esc(s: str) -> str:
    return s.replace("|", "\\|").replace("\n", " ").strip()


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def _col_to_idx(col: str) -> int:
    idx = 0
    for c in col.upper():
        idx = idx * 26 + (ord(c) - ord("A") + 1)
    return idx


def _walk_title(title_obj: Any) -> str | None:
    """Defensively extract title text from openpyxl Title rich-text structure."""
    if title_obj is None:
        return None
    try:
        tx = getattr(title_obj, "tx", None)
        if tx is None:
            return None
        rich = getattr(tx, "rich", None)
        if rich is None:
            return None
        parts = []
        paragraphs = getattr(rich, "paragraphs", None) or getattr(rich, "p", None) or []
        for p in paragraphs:
            runs = getattr(p, "r", None) or []
            for r in runs:
                t = getattr(r, "t", None)
                if t:
                    parts.append(t)
        text = "".join(parts).strip()
        return text or None
    except AttributeError:
        return None


def _find_used_range(ws) -> tuple[int, int, int, int] | None:
    """Return (min_row, min_col, max_row, max_col) 1-indexed inclusive, or None."""
    max_row = 0
    max_col = 0
    min_row: int | None = None
    min_col: int | None = None
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is None:
                continue
            if isinstance(v, str) and not v.strip():
                continue
            r, c = cell.row, cell.column
            if r > max_row:
                max_row = r
            if c > max_col:
                max_col = c
            if min_row is None or r < min_row:
                min_row = r
            if min_col is None or c < min_col:
                min_col = c
    if max_row == 0 or min_row is None or min_col is None:
        return None
    return (min_row, min_col, max_row, max_col)


def _build_data_table(ws, max_rows: int = SPREADSHEET_MAX_ROWS) -> str:
    bounds = _find_used_range(ws)
    if bounds is None:
        return ""
    min_row, min_col, max_row, max_col = bounds
    width = max_col - min_col + 1
    height = max_row - min_row + 1

    grid: list[list[str]] = [["" for _ in range(width)] for _ in range(height)]
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            grid[r - min_row][c - min_col] = _stringify(ws.cell(row=r, column=c).value)

    for merged in ws.merged_cells.ranges:
        anchor_val = _stringify(ws.cell(row=merged.min_row, column=merged.min_col).value)
        if not anchor_val:
            continue
        for r in range(merged.min_row, merged.max_row + 1):
            for c in range(merged.min_col, merged.max_col + 1):
                if min_row <= r <= max_row and min_col <= c <= max_col:
                    grid[r - min_row][c - min_col] = anchor_val

    original_data_rows = height - 1
    if original_data_rows > max_rows:
        grid = grid[: max_rows + 1]
        logger.warning(
            f"Sheet '{ws.title}' exceeds {max_rows} rows; truncating "
            f"(had {original_data_rows} data rows)"
        )

    if not grid:
        return ""

    header = [_esc(c) for c in grid[0]]
    body_rows = [[_esc(c) for c in row] for row in grid[1:]]
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * width) + " |",
    ]
    for row in body_rows:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)


def _chart_type(chart) -> str:
    cls = type(chart).__name__
    if cls in _CHART_TYPE_MAP:
        return _CHART_TYPE_MAP[cls]
    stripped = cls.removesuffix("Chart").removesuffix("3D")
    return stripped.lower() or cls.lower()


def _parse_range(ref: str) -> tuple[str, int, int, int, int] | None:
    if not ref or "[" in ref:
        return None
    m = _RANGE_RE.match(ref)
    if not m:
        return None
    sheet = m.group("quoted_sheet")
    if sheet is not None:
        sheet = sheet.replace("''", "'")
    else:
        sheet = m.group("sheet")
    col1 = _col_to_idx(m.group("col1"))
    row1 = int(m.group("row1"))
    col2 = _col_to_idx(m.group("col2")) if m.group("col2") else col1
    row2 = int(m.group("row2")) if m.group("row2") else row1
    return (sheet, row1, col1, row2, col2)


def _resolve_values(wb, ref: str) -> list[Any] | None:
    parsed = _parse_range(ref)
    if parsed is None:
        return None
    sheet_name, r1, c1, r2, c2 = parsed
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    values = []
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            values.append(ws.cell(row=r, column=c).value)
    return values


def _series_name(wb, series, default: str) -> str:
    try:
        tx = getattr(series, "tx", None)
        if tx is None:
            return default
        str_ref = getattr(tx, "strRef", None)
        if str_ref is not None and getattr(str_ref, "f", None):
            vals = _resolve_values(wb, str_ref.f)
            if vals and vals[0] is not None:
                return _stringify(vals[0])
        v = getattr(tx, "v", None)
        if v:
            return v
    except AttributeError:
        pass
    return default


def _series_data(wb, series) -> tuple[list[Any] | None, list[Any] | None, bool]:
    """Return (values, categories, is_external)."""
    values: list[Any] | None = None
    categories: list[Any] | None = None
    is_external = False
    try:
        val = getattr(series, "val", None)
        if val is not None:
            num_ref = getattr(val, "numRef", None)
            if num_ref is not None and getattr(num_ref, "f", None):
                values = _resolve_values(wb, num_ref.f)
                if values is None and "[" in num_ref.f:
                    is_external = True
            else:
                num_lit = getattr(val, "numLit", None)
                if num_lit is not None:
                    pts = getattr(num_lit, "pt", None) or []
                    values = [getattr(pt, "v", None) for pt in pts]
    except AttributeError:
        pass
    try:
        cat = getattr(series, "cat", None)
        if cat is not None:
            str_ref = getattr(cat, "strRef", None)
            num_ref = getattr(cat, "numRef", None)
            if str_ref is not None and getattr(str_ref, "f", None):
                categories = _resolve_values(wb, str_ref.f)
            elif num_ref is not None and getattr(num_ref, "f", None):
                categories = _resolve_values(wb, num_ref.f)
    except AttributeError:
        pass
    return values, categories, is_external


def _axis_label(chart, axis_attr: str) -> str | None:
    axis = getattr(chart, axis_attr, None)
    if axis is None:
        return None
    return _walk_title(getattr(axis, "title", None))


def _render_chart(wb, chart) -> str:
    title = _walk_title(getattr(chart, "title", None)) or "(untitled)"
    ctype = _chart_type(chart)
    lines = [f"### Chart: {title} ({ctype})"]

    x_label: str | None = None
    y_label: str | None = None
    z_label: str | None = None
    if ctype not in {"pie", "doughnut"}:
        x_label = _axis_label(chart, "x_axis")
        y_label = _axis_label(chart, "y_axis")
        z_label = _axis_label(chart, "z_axis")
        if x_label:
            lines.append(f"**X-axis:** {x_label}")
        if y_label:
            lines.append(f"**Y-axis:** {y_label}")
        if z_label:
            lines.append(f"**Z-axis:** {z_label}")

    series_list = list(getattr(chart, "series", None) or [])
    if not series_list:
        lines.append("")
        lines.append("_(no series)_")
        return "\n".join(lines)

    records = []  # (name, values, categories, is_external)
    for i, s in enumerate(series_list, 1):
        name = _series_name(wb, s, f"Series {i}")
        values, categories, is_external = _series_data(wb, s)
        records.append((name, values, categories, is_external))

    cat_keys = [tuple(_stringify(x) for x in (cs or [])) for _, _, cs, _ in records]
    all_have_data = all(v is not None for _, v, _, _ in records)
    all_have_cats = all(cs is not None and len(cs) > 0 for _, _, cs, _ in records)
    cats_equal = all_have_cats and len(set(cat_keys)) == 1
    can_fold = all_have_data and cats_equal and len(records) > 0

    lines.append("")
    if can_fold:
        cats = records[0][2] or []
        cat_header = x_label or "Category"
        header = [cat_header] + [name for name, _, _, _ in records]
        lines.append("| " + " | ".join(_esc(h) for h in header) + " |")
        lines.append("| " + " | ".join(["---"] * len(header)) + " |")
        for i, cat in enumerate(cats):
            row = [_stringify(cat)]
            for _, vals, _, _ in records:
                row.append(_stringify(vals[i]) if vals is not None and i < len(vals) else "")
            lines.append("| " + " | ".join(_esc(c) for c in row) + " |")
    else:
        for name, values, categories, is_external in records:
            lines.append(f"**Series: {name}**")
            if values is None:
                lines.append("_(external reference)_" if is_external else "_(no data)_")
                lines.append("")
                continue
            if categories and len(categories) == len(values):
                lines.append("| " + " | ".join(_esc(h) for h in ["Category", name]) + " |")
                lines.append("| --- | --- |")
                for cat, val in zip(categories, values):
                    lines.append(f"| {_esc(_stringify(cat))} | {_esc(_stringify(val))} |")
            else:
                lines.append(f"| {_esc(name)} |")
                lines.append("| --- |")
                for val in values:
                    lines.append(f"| {_esc(_stringify(val))} |")
            lines.append("")

    return "\n".join(lines).rstrip()


def _classify_sheet(ws) -> str:
    has_charts = bool(getattr(ws, "_charts", None))
    has_data = _find_used_range(ws) is not None
    if not has_data and not has_charts:
        return "empty"
    if has_charts and not has_data:
        return "chart"
    if has_data and has_charts:
        return "mixed"
    return "data"


def extract_spreadsheet(path: str, output_dir: str) -> str:
    """Extract a spreadsheet (.xlsx/.xls/.ods) to per-sheet markdown."""
    from openpyxl import load_workbook

    ext = os.path.splitext(path)[1].lower()
    work_path = path
    if ext != ".xlsx":
        work_path = _convert_to_xlsx(path, output_dir)

    try:
        wb = load_workbook(work_path, data_only=True)
    except Exception as e:
        raise RuntimeError(
            f"Failed to load workbook {os.path.basename(path)}: {e}"
        )

    sections: list[str] = []
    for ws in wb.worksheets:
        sections.append(f"## Sheet: {ws.title}")
        klass = _classify_sheet(ws)
        if klass == "empty":
            sections.append("_(empty sheet)_")
            continue
        if klass in ("data", "mixed"):
            table_md = _build_data_table(ws)
            if table_md:
                sections.append(table_md)
        if klass in ("chart", "mixed"):
            for chart in (ws._charts or []):
                sections.append(_render_chart(wb, chart))

    return "\n\n".join(sections)
